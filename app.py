import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import db_manager as dbm
from datetime import datetime
from tkinter import Tk, ttk, Frame, Label, Button, Entry, messagebox, Toplevel, Scrollbar, filedialog, Menu



root = Tk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()



class Main_Window():
    def __init__(self):
        self.root = root
        self.window()
        self.frame_window()
        self.filter_tipo = None
        self.order_name = "id"
        self.order_direction = "desc"
        self.ultima_coluna = None
        self.transaction_id = None
        self.head_values = ["↓", "Pessoa", "Tipo", "Categoria", "Descrição", "Local", "Valor", "Moeda", "Cotação", "Valor Total", "Data"]
        self.widgets_frame()
        self.form_aberto = False
        self.root.mainloop()

    def window(self):
        self.root.title("Transações")
        self.root.configure(background= "#8F8F8F")
        self.root.geometry(f"{int(screen_width*0.8)}x{int(screen_height * 0.8)}+{int(screen_width*0.1)}+{int(screen_height*0.08)}")
        self.root.resizable(False,False)
    
    def frame_window(self):
        self.frame_1 = Frame(self.root,bg="#D4D4D4",
                                bd=2, highlightbackground="#000000", highlightthickness=1)
        self.frame_1.place(relx=0.002,rely=0.005,relwidth=0.996,relheight=0.99)

    def carregar_transacoes(self):
        for item in self.lista_transacao.get_children():
            self.lista_transacao.delete(item)
            
        transacoes = dbm.todas_transacoes(tipo=self.filter_tipo, name=self.order_name, direct = self.order_direction)

        if transacoes:
            rend = 0
            desp = 0
            for transacao in transacoes:
                if transacao["Tipo"] == "Receita":
                    rend+= transacao["Valor_Total"]

                elif transacao["Tipo"] == "Despesa":
                    desp+= transacao["Valor_Total"]
            total = rend-desp
            total = 0.0 if total == 0 else total

            self.receita_button.config(text=f"Receita Total\n\nR$ {round(rend,2):.2f}")
            self.despesa_button.config(text=f"Despesa Total\n\nR$ {round(desp,2):.2f}")
            self.saldo_button.config(text=f"Saldo Total\n\nR$ {round(total,2):.2f}")
        else:
            self.receita_button.config(text=f"Receita Total\n\nR$ 0")
            self.despesa_button.config(text=f"Despesa Total\n\nR$ 0")
            self.saldo_button.config(text=f"Saldo Total\n\nR$ 0")

        self.lista_transacao.tag_configure("Receita", background="#f0fff1")
        self.lista_transacao.tag_configure("Despesa", background="#ffefef")

        for idx, transacao in enumerate(transacoes):
            if self.order_direction == "desc":
                index = len(transacoes)- idx
            elif self.order_direction == "asc":
                index = idx + 1

            tag = transacao["Tipo"]

            self.lista_transacao.insert("", "end", iid=transacao["id"],  text=str(index), values=(
                transacao["Pessoa"],
                transacao["Tipo"],
                transacao["Categoria"],
                transacao["Descricao"],
                transacao["Local"],
                transacao["Valor"],
                transacao["Moeda"],
                transacao["Cotacao"],
                transacao["Valor_Total"],
                transacao["Data"]
            ),
            tags=(tag,)
            )

    def selecionar_head(self, head_coluna):
        map_colunas = {
            "↓": "id",
            "↑": "id",
            "Pessoa": "Pessoa",
            "Tipo": "Tipo",
            "Categoria": "Categoria",
            "Descrição": "Descricao",
            "Local": "Local",
            "Valor": "Valor",
            "Moeda": "Moeda",
            "Cotação": "Cotacao",
            "Valor Total": "Valor_Total",
            "Data": "Data"
        }

        self.order_name = map_colunas[head_coluna]

        if self.ultima_coluna == map_colunas[head_coluna]:
            self.order_direction = "asc" if self.order_direction == "desc" else "desc"
        else:
            self.order_direction = "desc"

        if self.order_direction == "desc":
            self.head_values = ["↓", "Pessoa", "Tipo", "Categoria", "Descrição", 
                                "Local", "Valor", "Moeda", "Cotação","Valor Total", "Data"]
        elif self.order_direction == "asc":
            self.head_values = ["↑", "Pessoa", "Tipo", "Categoria", "Descrição", 
                                "Local", "Valor", "Moeda", "Cotação", "Valor Total", "Data"]

        self.ultima_coluna = map_colunas[head_coluna]
        self.widgets_frame()


    def exportar_excel(self):
        try:            
            arquivos = dbm.todas_transacoes(tipo=self.filter_tipo,
                                            name=self.order_name,
                                            direct=self.order_direction)
            if not arquivos:
                messagebox.showwarning("Aviso", "Não há transações para exportar.")
                return
            
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivo Excel", "*.xlsx")]
            )
            
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Transações"
            headers = list(arquivos[0].keys())[1:]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="002A5E", end_color="00224D", fill_type="solid")
            thin_border = Border(left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin"))

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            cor_escura = PatternFill(start_color="F5FAFF", end_color="F5FAFF", fill_type="solid")
            cor_clara = PatternFill(start_color="DFEFFF", end_color="DFEFFF", fill_type="solid")  

            for row_idx, arquivo in enumerate(arquivos, start=2):
                fill_cor = cor_escura if row_idx % 2 == 0 else cor_clara
                for col_idx, key in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=arquivo[key])
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                    cell.fill = fill_cor

            screen_width_excel = screen_width * 1.48
            proporcoes = (0.145, 0.06, 0.06, 0.338, 0.06, 0.06, 0.06, 0.06, 0.06, 0.06)
            proporcoes = [p / 1.023 for p in proporcoes]
            larguras_colunas = [screen_width_excel * p / 10 for p in proporcoes]

            for col_idx, largura in enumerate(larguras_colunas, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = largura

            wb.save(path)
            messagebox.showinfo("Sucesso", "Arquivo salvo com sucesso")

        except PermissionError:
            messagebox.showerror("Erro", "Erro: O arquivo está aberto.\nFeche para prosseguir")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado ao salvar: {e}")



    def widgets_frame(self):
        ###################################################| F U N Ç Õ E S |###################################################################################
        def abrir_form(dados = None):
            if not self.form_aberto:
                self.form_aberto = True
                Formulario(self.root, self, dados)
            
            elif self.form_aberto:
                messagebox.showinfo("Aviso", "Formulario está Aberto")

        def aplicar_filtro(tipo = None):
            self.filter_tipo = tipo
            self.carregar_transacoes()
        
        def formulario_de_edicao(event=None):
            if len(self.lista_transacao.selection()) != 1:
                messagebox.showerror("Erro", "Selecione apenas uma transação pra editar")
                return
            valores = self.lista_transacao.item(self.lista_transacao.selection())["values"]

            id = int(self.lista_transacao.selection()[0])
            lista_data = valores[9].split("/")[::-1]
            Data = lista_data[0] + '-' + lista_data[1] + '-' + lista_data[2] + ' 00:00:00.000000'
            dados = {   "Id":id,
                        "Pessoa": valores[0],
                        "Tipo": valores[1],
                        "Categoria": valores[2],
                        "Descricao": valores[3],
                        "Local": valores[4],
                        "Valor": valores[5],
                        "Moeda": valores[6],
                        "Cotacao": valores[7],
                        "Data": Data
                    }
            abrir_form(dados=dados)

        def deletar_transacao(event=None):
            ids = self.lista_transacao.selection()
            ids = [int(i) for i in ids]
            confirmar = messagebox.askyesno(
                "Confirmar Exclusão",
                f"Deseja excluir as transações selecionadas?"
            )

            if confirmar:
                mensagem = dbm.deletar_transacoes(ids=ids)
                self.carregar_transacoes()
                messagebox.showinfo("Transação",mensagem)
        

        def abrir_menu(event):
            row = self.lista_transacao.identify_row(event.y)
            if row:
                self.lista_transacao.selection_set(row)
            menu.tk_popup(event.x_root, event.y_root)

        ###################################################| L I S T A |#######################################################################################
        heads = self.head_values
        width = (0.04, 0.135, 0.06, 0.06, 0.308, 0.06, 0.06, 0.06, 0.06, 0.06, 0.06)
        self.lista_transacao = ttk.Treeview(self.frame_1,height=3,columns=heads)


        
        for idx,element in enumerate(heads):
            if idx == 0:
                position = "w"
            else:
                position = "center"

            self.lista_transacao.heading(f"#{idx}",text=element,anchor="center",command=lambda element=element: self.selecionar_head(element))
            self.lista_transacao.column(f"#{idx}", width=int(width[idx]*(screen_width*0.8)), anchor=position)

        self.lista_transacao.place(relx=0.01, rely=0.02, relwidth=0.97, relheight=0.78)
        self.lista_transacao_scrollbar = Scrollbar(self.frame_1, orient="vertical", bd=2, highlightbackground= "#000000")
        self.lista_transacao.configure(yscrollcommand=self.lista_transacao_scrollbar.set)
        self.lista_transacao_scrollbar.place(relx= 0.981, rely=0.021, relwidth=0.009, relheight=0.778)
        self.lista_transacao_scrollbar.config(command=self.lista_transacao.yview)
        self.lista_transacao.bind("<Double-1>", formulario_de_edicao)
        self.lista_transacao.bind("<Return>", formulario_de_edicao)
        self.lista_transacao.bind("<Delete>", deletar_transacao)

        menu = Menu(root, tearoff=0)
        menu.add_command(label="Editar", command=formulario_de_edicao)
        menu.add_command(label="Deletar", command=deletar_transacao)

        self.lista_transacao.bind("<Button-3>", abrir_menu)

        #################################################| B O T Õ E S |#########################################################################################
        # Adicionar Transação
        self.adicionar_transacao_button = Button(self.frame_1,text="Adicionar Transação", bd=3, command=abrir_form)
        self.adicionar_transacao_button.place(relx=0.7, rely= 0.825, relwidth=0.29, relheight= 0.075)

        # Exportar para Excel
        self.export_button = Button(self.frame_1,text="Exportar pra Excel", bd=3, command= self.exportar_excel)
        self.export_button.place(relx=0.7, rely= 0.9, relwidth=0.29, relheight= 0.075)

        # Receita Total
        self.receita_button = Button(self.frame_1, bd=2, background="#acffb0",activebackground="#7dff84",foreground="#003B04", font=("Arial", 10, "bold"), command= lambda: aplicar_filtro("Receita"))
        self.receita_button.place(relx=0.01, rely= 0.825, relwidth=0.2, relheight= 0.15)
        
        # Despesa Total
        self.despesa_button = Button(self.frame_1, bd=2, background="#ffa8a8", activebackground = "#ff8080", foreground="#3A0000", font=("Arial", 10, "bold"), command=lambda: aplicar_filtro("Despesa"))
        self.despesa_button.place(relx=0.22, rely= 0.825, relwidth=0.2, relheight= 0.15)
        
        # Saldo Total
        self.saldo_button = Button(self.frame_1, background="#aad6ff", activebackground="#82c0ff", foreground="#00203D", font=("Arial", 10, "bold"), command=lambda: aplicar_filtro(None))
        self.saldo_button.place(relx=0.43, rely= 0.825, relwidth=0.2, relheight= 0.15)

        self.carregar_transacoes()
        


class Formulario():
    
    def __init__(self,parent, main_window, dados = None):
        self.main_window = main_window
        self.root = Toplevel(parent)
        self.root.transient(parent)
        self.root.grab_set()
        self.root.focus_set()   
        self.root.protocol("WM_DELETE_WINDOW", self.fechar_form)
        self.dados = dados
        self.root.bind("<Escape>",lambda event:self.fechar_form())
        self.window()
        self.frame_window()
        self.widgets_frame()

    def window(self):
        self.root.title("Adicionar Transação")
        self.root.configure(background= "#8F8F8F")
        self.root.geometry(f"{int(screen_width * 0.2)}x{int(screen_height * 0.55)}+{int(screen_width * 0.4)}+{int(screen_height*0.225)}")
        self.root.resizable(False,False)

    def frame_window(self):
        self.frame_1 = Frame(self.root,bg="#D4D4D4",
                             bd=2, highlightbackground="#000000", highlightthickness=1)
        self.frame_1.place(relx=0.005,rely=0.005,relwidth=0.99,relheight=0.99)
    
    def fechar_form(self):
        self.main_window.form_aberto = False
        self.root.destroy()

    def proximo(self, event):
        event.widget.tk_focusNext().focus()
        return "break"


    def widgets_frame(self):
    ###########################################| F U N Ç Õ E S |#########################################################
        def combobox(parent, values, valor_inicial=None, readonly=True, x=0, y=0, width=0.1, height = 0.1):
            estado = "readonly" if readonly else "normal"
            combo = ttk.Combobox(parent, values=values, state=estado)
            
            if valor_inicial:
                combo.set(valor_inicial)
            
            combo.place(relx=x, rely=y, relwidth= width, relheight= height)
            return combo
        
        def validar_tamanho(entry,tam_max):
            if len(entry) > tam_max:
                return False
            return True

        def validar_pessoa(entry):
            if not validar_tamanho(entry=entry,tam_max=30):
                return False
            if not re.compile(r'^[A-Za-z\s]*$').match(entry):
                return False
            return True
        
        def validar_descricao(entry):
            if not validar_tamanho(entry=entry,tam_max=60):
                return False
            if not re.compile(r'^[A-Za-z0-9\s]*$').match(entry):
                return False
            return True
        
        def validar_valor(entry):
            entry = entry.replace(',','.')
            if not re.compile(r'^\d{0,6}(\.\d{0,2})?$').match(entry):
                return False
            return True
        
        def validar_cotacao(entry):
            entry = entry.replace(',','.')
            if not re.compile(r'^\d{0,2}(\.\d{0,4})?$').match(entry):
                return False
            return True
                        
        def atualizar_posicoes_tipo(event):
            if self.entry_tipo.get() != "":
                self.pos_categoria = self.pos_tipo + 1
                self.label_categoria.place(relx=0.5, rely=0.025 + self.pos_categoria * 0.1, anchor="center")
                self.entry_categoria.place(relx=0.01, rely=0.05 + self.pos_categoria * 0.1, relwidth=0.98, relheight=0.05)

                self.pos_descricao = self.pos_categoria + 1
                self.label_descricao.place(relx=0.5, rely=0.025 + self.pos_descricao *0.1, anchor="center")
                self.entry_descricao.place(relx=0.01, rely=0.05 + self.pos_descricao * 0.1, relwidth=0.98, relheight=0.05)

                self.pos_local = self.pos_descricao + 1
                self.label_local.place(relx=0.5, rely=0.025 + self.pos_local *0.1, anchor="center")
                self.entry_local.place(relx=0.01, rely=0.05 + self.pos_local *0.1, relwidth=0.98, relheight=0.05)

                self.pos_valor = self.pos_local + 1
                self.label_valor.place(relx=0.5, rely=0.025 + self.pos_valor *0.1, anchor="center")
                self.entry_valor.place(relx=0.01,rely=0.05 + self.pos_valor * 0.1,relwidth=0.59,relheight=0.05)
                self.entry_moeda.place(relx=0.6, rely=0.05 + self.pos_valor *0.1, relwidth=0.39, relheight=0.05)
                
                self.pos_cotacao = self.pos_valor + 1
                self.label_cotacao.place(relx=0.5, rely=0.025 + self.pos_cotacao *0.1, anchor="center")
                self.entry_cotacao.place(relx=0.01, rely=0.05 + self.pos_cotacao*0.1, relwidth=0.98, relheight=0.05)
                if self.entry_moeda.get() == "Real":
                    self.pos_cotacao = self.pos_valor
                    self.label_cotacao.place_forget()
                    self.entry_cotacao.place_forget()
                    
                self.pos_data = self.pos_cotacao + 1
                self.label_data.place(relx=0.5, rely=0.025 + self.pos_data *0.1, anchor="center")
                self.entry_data_dia.place(relx=0.29, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                self.entry_data_mes.place(relx=0.43, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                self.entry_data_ano.place(relx=0.57, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
            
            if self.entry_tipo.get() == "Receita":
                self.entry_categoria.set("")
                self.entry_categoria["values"] = ('Salário', 'Rendimento', 'Outro')
            
            elif self.entry_tipo.get() == "Despesa":
                self.entry_categoria.set("")
                self.entry_categoria['values'] = ('Alimentação', 'Transporte', 'Lazer', 'Saúde', 'Hospedagem', 'Passagem', 'Documento', 'Outro')

        def atualizar_posicoes_moeda(event):
            if self.entry_moeda.get() != "Real":
                self.pos_cotacao = self.pos_valor + 1

                self.label_cotacao.place(relx=0.5, rely=0.025 + self.pos_cotacao*0.1, anchor="center")
                self.entry_cotacao.place(relx=0.01, rely=0.05 + self.pos_cotacao*0.1, relwidth=0.98, relheight=0.05)

                self.pos_data = self.pos_cotacao + 1
                self.label_data.place(relx=0.5, rely=0.025 + self.pos_data*0.1, anchor="center")
                self.entry_data_dia.place(relx=0.29, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                self.entry_data_mes.place(relx=0.43, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                self.entry_data_ano.place(relx=0.57, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)

            elif self.entry_moeda.get() == "Real":
                self.label_cotacao.place_forget()
                self.entry_cotacao.place_forget()

                self.pos_data = self.pos_valor + 1

                self.label_data.place(relx=0.5, rely=0.025 + self.pos_data *0.1, anchor="center")
                self.entry_data_dia.place(relx=0.29, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                self.entry_data_mes.place(relx=0.43, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                self.entry_data_ano.place(relx=0.57, rely=0.05 + self.pos_data *0.1, relwidth=0.14, relheight=0.05)
                
        def salvar_form():
            pessoa = self.entry_pessoa.get()
            tipo = self.entry_tipo.get()
            categoria = self.entry_categoria.get()
            descricao = self.entry_descricao.get()
            local = self.entry_local.get()
            valor = self.entry_valor.get().replace(",",".")
            moeda = self.entry_moeda.get()
            cotacao = "1" if "Real" in self.entry_moeda.get() else self.entry_cotacao.get().replace(",",".")
            data = datetime(
                            year=int(self.entry_data_ano.get()),
                            month=int(self.entry_data_mes.get()),
                            day=int(self.entry_data_dia.get())
                            )
            
            campos = {
            "Pessoa": pessoa,
            "Tipo": tipo,
            "Categoria": categoria,
            "Descricao": descricao,
            "Local": local,
            "Valor": valor,
            "Moeda": moeda,
            "Cotacao": cotacao,
            "Data": data
            }

            for nome, valor_campo in campos.items():
                if nome == "Descricao":
                    continue
                elif not valor_campo:
                    messagebox.showerror("Erro", f'O campo "{nome}" precisa ser preenchido!')
                    return
            
            if float(str(campos["Valor"]).replace(",",".")) == 0:
                messagebox.showerror("Erro", 'O campo "Valor" deve ser maior que 0')
                return

            if float(str(campos["Cotacao"]).replace(",",".")) == 0:
                messagebox.showerror("Erro", 'O campo "Cotação" deve ser maior que 0')
                return

            form = {
                    "Pessoa": str(pessoa).strip(),
                    "Tipo": str(tipo),
                    "Categoria": str(categoria),
                    "Descricao": str(descricao).strip(),
                    "Local": str(local),
                    "Valor": float(valor),
                    "Moeda": str(moeda),
                    "Cotacao": float(cotacao),
                    "Data": data
                    }
            
            
            mensagem = dbm.adicionar_transacao(form)
            self.fechar_form()
            self.main_window.carregar_transacoes()
            messagebox.showinfo("Transação",mensagem)


        def atualizar_form(id):
            pessoa = self.entry_pessoa.get()
            tipo = self.entry_tipo.get()
            categoria = self.entry_categoria.get()
            descricao = self.entry_descricao.get()
            local = self.entry_local.get()
            valor = self.entry_valor.get().replace(",",".")
            moeda = self.entry_moeda.get()
            cotacao = "1" if "Real" in self.entry_moeda.get() else self.entry_cotacao.get().replace(",",".")
            data = datetime(
                            year=int(self.entry_data_ano.get()),
                            month=int(self.entry_data_mes.get()),
                            day=int(self.entry_data_dia.get())
                            )
            
            campos = {
            "Id":id,
            "Pessoa": pessoa,
            "Tipo": tipo,
            "Categoria": categoria,
            "Descricao": descricao,
            "Local": local,
            "Valor": valor,
            "Moeda": moeda,
            "Cotacao": cotacao,
            "Data": data
            }

            for nome, valor_campo in campos.items():
                if nome == "Descricao":
                    continue
                elif not valor_campo:
                    messagebox.showerror("Erro", f"O campo '{nome}' precisa ser preenchido!")
                    return
            
            if float(str(campos["Valor"]).replace(",",".")) == 0:
                messagebox.showerror("Erro", 'O campo "Valor" deve ser maior que 0')
                return
            if float(str(campos["Cotacao"]).replace(",",".")) == 0:
                messagebox.showerror("Erro", 'O campo "Cotação" deve ser maior que 0')
                return

            form = {"Id": int(id),
                    "Pessoa": str(pessoa).strip(),
                    "Tipo": str(tipo),
                    "Categoria": str(categoria),
                    "Descricao": str(descricao).strip(),
                    "Local": str(local),
                    "Valor": float(valor),
                    "Moeda": str(moeda),
                    "Cotacao": float(cotacao),
                    "Data": data
                    }
            mensagem = dbm.update_transacao(form = form)
            self.fechar_form()
            self.main_window.carregar_transacoes()
            messagebox.showinfo("Transação",mensagem)



    ########################################| L A B E L S |##########################################################       
        self.pos_pessoa = 0
        self.pos_tipo = None
        self.pos_categoria = None
        self.pos_descricao = None
        self.pos_local = None
        self.pos_valor = None
        self.pos_data = None
        self.pos_cotacao = None

        # Pessoa
        self.label_pessoa = Label(self.frame_1, text= "Pessoa")
        self.label_pessoa.place(relx = 0.5, rely = 0.025 + self.pos_pessoa * 0.1, anchor="center",)

        vcmd_pessoa = (self.frame_1.register(validar_pessoa), "%P")
        self.entry_pessoa = Entry(self.frame_1, validate="key", validatecommand=vcmd_pessoa)
        self.entry_pessoa.place(relx=0.01,rely= 0.05 + self.pos_pessoa * 0.1, relwidth=0.98,relheight=0.05)

        # Tipo
        self.pos_tipo = self.pos_pessoa + 1

        self.label_tipo = Label(self.frame_1, text= "Tipo")
        self.label_tipo.place(relx = 0.5, rely = 0.025 + self.pos_tipo * 0.1, anchor="center",)

        self.entry_tipo = combobox(parent=self.frame_1,values=("Receita","Despesa"),readonly=True,x=0.01,y=0.05 + self.pos_tipo * 0.1,
                                                        width=0.98, height=0.05)
        self.lista_categoria = tuple()
        
        self.entry_tipo.bind("<<ComboboxSelected>>", atualizar_posicoes_tipo)

        #Categoria
        self.pos_categoria = self.pos_tipo + 1
        self.label_categoria = Label(self.frame_1, text= "Categoria")
        self.label_categoria.place(relx = 0.5, rely = 0.025 + self.pos_categoria * 0.1, anchor="center")
        self.label_categoria.place_forget()

        self.entry_categoria = combobox(parent=self.frame_1,values=self.lista_categoria,readonly=True,x=0.01,y=0.05 + self.pos_categoria * 0.1,
                                                        width=0.98, height=0.05)
        self.entry_categoria.place_forget()

        # Descrição
        self.pos_descricao = self.pos_tipo + 1

        self.label_descricao = Label(self.frame_1, text= "Descrição")
        self.label_descricao.place(relx = 0.5, rely = 0.025 + self.pos_descricao * 0.1, anchor="center")

        vcmd_descricao = (self.frame_1.register(validar_descricao), "%P")
        self.entry_descricao = Entry(self.frame_1, validate="key", validatecommand=vcmd_descricao)
        self.entry_descricao.place(relx=0.01,rely=0.05 + self.pos_descricao * 0.1, relwidth=0.98,relheight=0.05)

        # Local
        self.pos_local = self.pos_descricao + 1

        self.label_local = Label(self.frame_1, text= "Local")
        self.label_local.place(relx = 0.5, rely = 0.025 + self.pos_local * 0.1, anchor="center",)

        self.entry_local = combobox(parent=self.frame_1,values=("Brasil","Coreia","Japão","Outro"),readonly=True,x=0.01,y=0.05 + self.pos_local * 0.1,
                                                        width=0.98, height=0.05)

        # Valor
        self.pos_valor = self.pos_local + 1

        self.label_valor = Label(self.frame_1, text= "Valor")
        self.label_valor.place(relx = 0.5, rely = 0.025 + self.pos_valor * 0.1, anchor="center")    
        
        vcmd_valor = (self.frame_1.register(validar_valor), "%P")
        self.entry_valor = Entry(self.frame_1, validate="key", validatecommand=vcmd_valor)
        self.entry_valor.place(relx=0.01,rely=0.05 + self.pos_valor * 0.1,relwidth=0.59,relheight=0.05)

        # Moeda
        self.entry_moeda = combobox(parent=self.frame_1,values=("Real","Dolar","Won","Yen"),
                                            readonly=True,x=0.6,y=0.05 + self.pos_valor * 0.1,
                                            valor_inicial= "Real", width=0.39, height=0.05)
        
        self.entry_moeda.bind("<<ComboboxSelected>>", atualizar_posicoes_moeda)

        # Cotação
        self.pos_cotacao = self.pos_valor
        self.label_cotacao = Label(self.frame_1, text="Cotação")
        self.label_cotacao.place(relx = 0.5, rely = 0.025 + self.pos_cotacao * 0.1, anchor="center")
        self.label_cotacao.place_forget()

        vcmd_cotacao = (self.frame_1.register(validar_cotacao), "%P")
        self.entry_cotacao = Entry(self.frame_1, validate="key", validatecommand=vcmd_cotacao)
        self.entry_cotacao.place(relx=0.01,rely= 0.05 + self.pos_cotacao * 0.1, relwidth=0.98,relheight=0.05)
        self.entry_cotacao.place_forget()
        
        # Data
        self.pos_data = self.pos_cotacao + 1 
        self.label_data = Label(self.frame_1, text= "Data")
        self.label_data.place(relx = 0.5, rely = 0.025 + self.pos_data * 0.1, anchor="center")
        dias = [i for i in range(1,32)]
        meses = [i for i in range(1,13)]
        anos = [i for i in range(1900,2101)]
        hoje = datetime.today()

        self.entry_data_dia = combobox(parent=self.frame_1,values=dias, valor_inicial=hoje.day,
                                            readonly=True,x=0.29,y=0.05 + self.pos_data * 0.1, width=0.14, height=0.05)
        self.entry_data_mes = combobox(parent=self.frame_1,values=meses, valor_inicial= hoje.month,
                                            readonly=True,x=0.43,y=0.05 + self.pos_data * 0.1, width=0.14, height=0.05)
        self.entry_data_ano = combobox(parent=self.frame_1,values=anos, valor_inicial= hoje.year,
                                            readonly=True,x=0.57,y=0.05 + self.pos_data * 0.1, width=0.14, height=0.05)
        
        self.entry_pessoa.bind("<Return>", self.proximo)
        self.entry_tipo.bind("<Return>", self.proximo)
        self.entry_categoria.bind("<Return>", self.proximo)
        self.entry_descricao.bind("<Return>", self.proximo)
        self.entry_local.bind("<Return>", self.proximo)
        self.entry_valor.bind("<Return>", self.proximo)
        self.entry_moeda.bind("<Return>", self.proximo)
        self.entry_cotacao.bind("<Return>", self.proximo)
        self.entry_data_dia.bind("<Return>", self.proximo)
        self.entry_data_mes.bind("<Return>", self.proximo)
        self.entry_data_ano.bind("<Return>", lambda event: self.save_button.focus_set())
    
        

    ########################################| B O T Õ E S |###################################################
        # Salvar
        self.save_button = Button(self.frame_1,text="Salvar", command=salvar_form, bd=3)
        self.save_button.place(relx=0, rely= 0.925, relwidth=0.5, relheight= 0.075)
        self.save_button.bind("<Return>", lambda event: self.save_button.invoke())

        # Cancelar
        self.close_button = Button(self.frame_1,text="Cancelar", command= self.fechar_form, bd=3)
        self.close_button.place(relx=0.5, rely= 0.925, relwidth=0.5, relheight= 0.075)
        

        if self.dados:

            self.root.title("Atualizar Transação")
            data_raw = self.dados["Data"]
            data = (data_raw.split()[0]).split("-")

            self.entry_pessoa.insert(0, self.dados["Pessoa"])
            self.entry_tipo.set(self.dados["Tipo"])
            self.entry_descricao.insert(0,self.dados["Descricao"])
            self.entry_local.set(self.dados["Local"])
            self.entry_valor.insert(0, str(self.dados["Valor"]))
            self.entry_moeda.set(self.dados["Moeda"])
            self.entry_cotacao.insert(0, str(self.dados["Cotacao"]))
            self.entry_data_dia.set(data[2])
            self.entry_data_mes.set(data[1])
            self.entry_data_ano.set(data[0])
            atualizar_posicoes_tipo(None)
            atualizar_posicoes_moeda(None)
            self.entry_categoria.set(self.dados["Categoria"])
            
            if self.entry_moeda.get() == "Real":
                self.label_cotacao.place_forget()
                self.entry_cotacao.place_forget()


            self.save_button = Button(self.frame_1,text="Atualizar", command=lambda:atualizar_form(self.dados["Id"]), bd=3)
            self.save_button.place(relx=0, rely= 0.925, relwidth=0.5, relheight= 0.075)
            self.save_button.bind("<Return>", lambda event: self.save_button.invoke())

            
            


if __name__ == "__main__":
    main_window = Main_Window()
